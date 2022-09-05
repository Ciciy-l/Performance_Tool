# -*- coding: UTF-8 -*-
import os

from openpyxl import load_workbook


class ExcelOperations:

    def __init__(self, excel_path):
        self.excel_path = os.path.abspath(excel_path)
        self.wb = load_workbook(self.excel_path)

    def get_sheet_list(self, list_type="name"):
        """
        获取sheet名称或对象列表
        :param list_type:"name"返回名称列表，"obj"返回对象列表
        """
        sheet_list = []
        if list_type == "name":
            for sheet in self.wb.sheetnames:
                sheet_list.append(sheet)
        elif list_type == "obj":
            for sheet in self.wb:
                sheet_list.append(sheet)
        return sheet_list

    def get_sheet_data(self, sheet_name: str):
        """
        传入sheet名称，将每行的数据以字典的形式获取，并放在字典中返回
        :param sheet_name: sheet名称
        """
        sheet_obj = self.wb[sheet_name]
        key_tuple = sheet_obj[1]
        key_list = []
        sheet_data_list = []
        [key_list.append(key_cell.value) for key_cell in key_tuple]
        for row in sheet_obj.iter_rows(min_row=2, max_row=sheet_obj.max_row):
            value_list = []
            data_dict = {}
            [value_list.append(cell.value) for cell in row]
            [data_dict.update({key: value}) for key, value in zip(key_list, value_list)]
            sheet_data_list.append(data_dict)
        return sheet_data_list

    def insert_blank_line(self, position: int, rows: int):
        """
        在指定行之前插入若干行
        :param position: 指定行号
        :param rows: 插入行数目
        """
        current_sheet = self.wb.active
        [current_sheet.insert_rows(position) for _ in range(rows)]

    def save_wb(self, file_path=None):
        """
        保存修改，默认覆盖保存，也可指定保存路径
        """
        file_path = self.excel_path if not file_path else file_path
        self.wb.save(file_path)
        self.wb.close()

    def close_wb(self):
        """
        关闭工作簿
        """
        self.wb.close()

    def switch_sheet(self, sheet_name: str):
        sheet_obj = self.wb[sheet_name]
        return sheet_obj
